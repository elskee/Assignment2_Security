"""
Excel Handler Module
Handles reading and writing vulnerability data from/to Excel files.
"""
import openpyxl
from typing import List, Dict, Any
import os


class ExcelHandler:
    """Handles Excel file operations for vulnerability data."""
    
    def __init__(self, file_path: str):
        """
        Initialize the Excel handler.
        
        Args:
            file_path: Path to the Excel file
        """
        self.file_path = file_path
        self.workbook = None
        self.worksheet = None
        self.headers = []
        self.data = []
        self.results_worksheet = None  # New worksheet for detailed results
        self.results_worksheet_initialized = False
        
    def load(self) -> List[Dict[str, Any]]:
        """
        Load data from the Excel file.
        
        Returns:
            List of dictionaries containing vulnerability data
        """
        if not os.path.exists(self.file_path):
            raise FileNotFoundError(f"Excel file not found: {self.file_path}")
        
        self.workbook = openpyxl.load_workbook(self.file_path)
        # Always use Sheet1 for source data, not the active sheet
        self.worksheet = self.workbook['Sheet1'] if 'Sheet1' in self.workbook.sheetnames else self.workbook.active
        
        # Read headers from first row
        self.headers = []
        for cell in self.worksheet[1]:
            self.headers.append(cell.value if cell.value else "")
        
        # Read data rows
        self.data = []
        for row_idx, row in enumerate(self.worksheet.iter_rows(min_row=2, values_only=False), start=2):
            row_data = {
                'row_index': row_idx,
                'data': {}
            }
            for col_idx, (header, cell) in enumerate(zip(self.headers, row)):
                row_data['data'][header] = cell.value if cell.value else ""
            
            # Only include rows that have some data
            if any(row_data['data'].values()):
                self.data.append(row_data)
        
        return self.data
    
    def update_found_in_projects(self, row_index: int, results: List[Dict[str, Any]]) -> None:
        """
        Update the "Found in other projects" column for a specific row.
        
        Args:
            row_index: Excel row index (1-based)
            results: List of repository results with 'repo_name', 'stars', and 'url'
        """
        if not self.worksheet:
            raise RuntimeError("Excel file not loaded. Call load() first.")
        
        # Find the column index for "Found in other projects"
        found_col_idx = None
        for idx, header in enumerate(self.headers, start=1):
            if header == "Found in other projects":
                found_col_idx = idx
                break
        
        if found_col_idx is None:
            raise ValueError("Column 'Found in other projects' not found in Excel file")
        
        # Format results as newline-separated entries
        formatted_results = []
        for result in results:
            repo_name = result.get('repo_name', 'Unknown')
            stars = result.get('stars', 0)
            url = result.get('url', '')
            formatted_results.append(f"{repo_name} ({stars} stars) | {url}")
        
        result_text = "\n".join(formatted_results)
        
        # Update the cell
        self.worksheet.cell(row=row_index, column=found_col_idx, value=result_text)
    
    def update_found_in_projects_none(self, row_index: int) -> None:
        """
        Update the "Found in other projects" column to indicate no results were found.
        
        Args:
            row_index: Excel row index (1-based)
        """
        if not self.worksheet:
            raise RuntimeError("Excel file not loaded. Call load() first.")
        
        # Find the column index for "Found in other projects"
        found_col_idx = None
        for idx, header in enumerate(self.headers, start=1):
            if header == "Found in other projects":
                found_col_idx = idx
                break
        
        if found_col_idx is None:
            raise ValueError("Column 'Found in other projects' not found in Excel file")
        
        # Update the cell with "None found"
        self.worksheet.cell(row=row_index, column=found_col_idx, value="None found")
    
    def save(self, backup: bool = True) -> None:
        """
        Save the Excel file.
        
        Args:
            backup: If True, create a backup of the original file
        """
        if not self.workbook:
            raise RuntimeError("Excel file not loaded. Call load() first.")
        
        # Create backup if requested
        if backup and os.path.exists(self.file_path):
            backup_path = f"{self.file_path}.bak"
            if os.path.exists(backup_path):
                os.remove(backup_path)
            import shutil
            shutil.copy2(self.file_path, backup_path)
        
        # Save the workbook
        self.workbook.save(self.file_path)
    
    def close(self) -> None:
        """Close the Excel workbook."""
        if self.workbook:
            self.workbook.close()
            self.workbook = None
            self.worksheet = None
            self.results_worksheet = None
    
    def get_vulnerability_data(self) -> List[Dict[str, Any]]:
        """
        Get vulnerability data from loaded Excel.
        
        Returns:
            List of dictionaries with vulnerability information
        """
        return self.data
    
    def _initialize_results_worksheet(self) -> None:
        """
        Initialize the 'Search Results' worksheet with headers.
        Creates the worksheet if it doesn't exist.
        """
        if not self.workbook:
            raise RuntimeError("Excel file not loaded. Call load() first.")
        
        # Check if 'Search Results' worksheet already exists
        if "Search Results" in self.workbook.sheetnames:
            self.results_worksheet = self.workbook["Search Results"]
        else:
            # Create new worksheet
            self.results_worksheet = self.workbook.create_sheet("Search Results")
            
            # Add headers
            headers = [
                "CVE ID",
                "Vulnerability Type",
                "Repository Name",
                "Stars",
                "Repository URL",
                "Commit URL",
                "File Path",
                "Code Snippet"
            ]
            
            for col_idx, header in enumerate(headers, start=1):
                cell = self.results_worksheet.cell(row=1, column=col_idx, value=header)
                # Make header bold
                from openpyxl.styles import Font
                cell.font = Font(bold=True)
        
        self.results_worksheet_initialized = True
    
    def add_result_to_worksheet(self, cve_id: str, vulnerability_type: str, 
                                result: Dict[str, Any]) -> None:
        """
        Add a single result to the 'Search Results' worksheet.
        
        Args:
            cve_id: CVE identifier
            vulnerability_type: Type of vulnerability
            result: Dictionary containing result data (repo_name, stars, url, 
                   commit_url, file_path, code_snippet)
        """
        if not self.results_worksheet_initialized:
            self._initialize_results_worksheet()
        
        # Find the next empty row
        next_row = self.results_worksheet.max_row + 1
        
        # Add data to the row
        self.results_worksheet.cell(row=next_row, column=1, value=cve_id)
        self.results_worksheet.cell(row=next_row, column=2, value=vulnerability_type)
        self.results_worksheet.cell(row=next_row, column=3, value=result.get('repo_name', ''))
        self.results_worksheet.cell(row=next_row, column=4, value=result.get('stars', 0))
        self.results_worksheet.cell(row=next_row, column=5, value=result.get('url', ''))
        self.results_worksheet.cell(row=next_row, column=6, value=result.get('commit_url', ''))
        self.results_worksheet.cell(row=next_row, column=7, value=result.get('file_path', ''))
        
        # Truncate code snippet if too long (Excel cell limit is 32,767 characters)
        code_snippet = result.get('code_snippet', '')
        if len(code_snippet) > 30000:
            code_snippet = code_snippet[:30000] + "\n... (truncated)"
        self.results_worksheet.cell(row=next_row, column=8, value=code_snippet)
    
    def __enter__(self):
        """Context manager entry."""
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager exit."""
        self.close()

